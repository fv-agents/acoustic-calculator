#!/usr/bin/env python3
"""
Lumenear Akoestische Calculator V4 — Complete Build Script
===========================================================
Bouwt het volledige Excel-bestand met 4 sheets:
  1. Calculator (interface)
  2. Materiaaldata (absorptiewaarden, presets)
  3. Productdata (89 producten)
  4. Peutz Testdata (EN-ISO 354)

Vereisten:
  pip install openpyxl pandas

Gebruik:
  python build_lumenear_calculator_v4.py
  python build_lumenear_calculator_v4.py --csv /pad/naar/lumenear_2026_acoustic_data.csv
  python build_lumenear_calculator_v4.py --output /pad/naar/output.xlsx

CSV moet semicolon-separated zijn met kolommen:
  Product_Family;Product_Variant;Diameter_mm;Acoustic_Surface_m2;
  aw_ISO11654;a_125Hz;a_250Hz;a_500Hz;a_1000Hz;a_2000Hz;a_4000Hz;Mounting_Type
"""

import argparse
import math
import os
import sys

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("FOUT: openpyxl niet geïnstalleerd. Run: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("FOUT: pandas niet geïnstalleerd. Run: pip install pandas")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
# CONFIGURATIE & STYLES
# ══════════════════════════════════════════════════════════════

DARK = "1B2A4A"
MED = "2F5496"
LGRAY = "F5F5F5"
WHITE = "FFFFFF"
INPUT_BG = "DBEAFE"
INPUT_C = "1E40AF"
GREEN_BG = "DCFCE7"
RED_BG = "FEE2E2"
ORANGE_BG = "FFF7ED"
SPEECH_BG = "FEF9C3"
BORDER_C = "D1D5DB"

tf = Font(bold=True, size=16, color=DARK, name="Calibri")
sf = Font(bold=True, size=13, color=MED, name="Calibri")
lf = Font(size=11, name="Calibri")
lbf = Font(bold=True, size=11, name="Calibri")
inp_f = Font(size=11, color=INPUT_C, bold=True, name="Calibri")
inp_fill = PatternFill("solid", fgColor=INPUT_BG)
df_ = Font(size=10, name="Calibri")
hf = Font(size=9, italic=True, color="9CA3AF", name="Calibri")
hdr_f = Font(bold=True, size=10, color=WHITE, name="Calibri")
hdr_fill = PatternFill("solid", fgColor=DARK)
sub_fill = PatternFill("solid", fgColor="E8EDF5")
grn_fill = PatternFill("solid", fgColor=GREEN_BG)
red_fill = PatternFill("solid", fgColor=RED_BG)
orng_fill = PatternFill("solid", fgColor=ORANGE_BG)
bdr = Border(
    left=Side("thin", BORDER_C), right=Side("thin", BORDER_C),
    top=Side("thin", BORDER_C), bottom=Side("thin", BORDER_C),
)
bdr2 = Border(
    left=Side("medium", DARK), right=Side("medium", DARK),
    top=Side("medium", DARK), bottom=Side("medium", DARK),
)


def shdr(ws, row, c1, c2):
    """Style header row."""
    for c in range(c1, c2 + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = hdr_f
        cl.fill = hdr_fill
        cl.border = bdr
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def srange(ws, r1, r2, c1, c2):
    """Style data range with alternating rows."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cl = ws.cell(row=r, column=c)
            cl.font = df_
            cl.border = bdr
            if r % 2 == 0:
                cl.fill = PatternFill("solid", fgColor=LGRAY)


# ══════════════════════════════════════════════════════════════
# MATERIAALDATA — BRONGEBASEERD
# Bronnen: Acoustic.ua, Sengpiel Audio, Commercial Acoustics
# ══════════════════════════════════════════════════════════════

#                                          125   250   500   1k    2k    4k
FLOORS = [
    ("Weet ik niet",                       .01,  .01,  .02,  .02,  .02,  .03),
    ("Harde vloer (beton/tegels)",         .01,  .01,  .02,  .02,  .02,  .03),
    ("Linoleum / vinyl / PVC",             .02,  .02,  .03,  .04,  .04,  .05),
    ("Parquet / laminaat",                 .04,  .04,  .07,  .06,  .06,  .07),
    ("Dun tapijt",                         .02,  .06,  .14,  .37,  .60,  .65),
    ("Dik tapijt (+ ondervloer)",          .08,  .24,  .57,  .69,  .71,  .73),
]

WALLS = [
    ("Beton / steen",                      .01,  .01,  .02,  .02,  .02,  .03),
    ("Tegels / natuursteen",               .01,  .01,  .01,  .01,  .02,  .02),
    ("Stucwerk",                           .02,  .02,  .03,  .04,  .04,  .04),
    ("Gipsplaat (standaard)",              .29,  .10,  .05,  .04,  .07,  .09),
    ("Gipsplaat + isolatie",               .30,  .12,  .08,  .06,  .06,  .05),
    ("Houten panelen",                     .15,  .20,  .10,  .10,  .05,  .05),
    ("Gips + weinig glas (~25%)",          .24,  .09,  .05,  .04,  .06,  .07),
    ("Gips + veel glas (~50%)",            .20,  .08,  .05,  .04,  .05,  .06),
    ("Overwegend glas (~75%)",             .15,  .07,  .04,  .03,  .03,  .04),
    ("Glazen systeemwand",                 .10,  .06,  .04,  .03,  .02,  .02),
    ("Zachte wandbekleding",               .05,  .15,  .25,  .35,  .30,  .25),
    ("Akoestische wandpanelen",            .10,  .20,  .40,  .50,  .45,  .40),
]

CEILINGS = [
    ("Weet ik niet",                       .01,  .01,  .02,  .02,  .02,  .03),
    ("Betonplafond",                       .01,  .01,  .02,  .02,  .02,  .03),
    ("Gipsplafond (dicht)",                .15,  .10,  .06,  .04,  .04,  .05),
    ("Houten plafond",                     .15,  .20,  .10,  .10,  .05,  .05),
    ("Metalen plafond (dicht)",            .15,  .10,  .08,  .06,  .05,  .04),
    ("Metalen plafond (geperforeerd)",     .15,  .25,  .40,  .55,  .50,  .45),
    ("Open plafond (installaties)",        .10,  .10,  .08,  .06,  .05,  .05),
    ("Systeemplafond (oud/dun)",           .05,  .15,  .30,  .45,  .50,  .45),
    ("Systeemplafond (standaard)",         .10,  .25,  .55,  .70,  .70,  .65),
    ("Systeemplafond (hoog absorberend)",  .30,  .65,  .90,  .95,  .90,  .85),
]

FURNITURE = [
    ("Weet ik niet",                .005),
    ("Leeg / weinig meubels",       .005),
    ("Kantoor - bureaus+stoelen",   .08),
    ("Kantoor - volledig ingericht", .10),
    ("Vergaderruimte",              .06),
    ("Restaurant / horeca",         .05),
    ("Lounge / gestoffeerd",        .12),
    ("Retail / winkel",             .03),
    ("Onderwijs / school",          .06),
    ("Zorg / medisch",              .04),
]

OTHER_AC = [
    ("1. Niks extra",               .00),
    ("2. Alleen planten",           .01),
    ("3. Dunne gordijnen",          .02),
    ("4. Losse tapijten",           .03),
    ("5. Raamdecoratie",            .04),
    ("6. Zware gordijnen",          .05),
    ("7. Bureauschermen",           .06),
    ("8. Gordijnen + tapijten",     .07),
    ("9. Schermen + gordijnen",     .09),
    ("10. Schermen + panelen",      .11),
    ("11. Roomdividers",            .14),
    ("12. Compleet pakket",         .18),
]

PRESETS = [
    ("(zelf invullen)",       "",                           "",                         "",                                  "",                              "",                          0.8),
    ("Kantoor (open plan)",   "Linoleum / vinyl / PVC",    "Gipsplaat + isolatie",     "Systeemplafond (standaard)",        "Kantoor - bureaus+stoelen",     "7. Bureauschermen",         0.6),
    ("Kantoor (celkantoor)",  "Dun tapijt",                "Gipsplaat + isolatie",     "Systeemplafond (standaard)",        "Kantoor - bureaus+stoelen",     "2. Alleen planten",         0.8),
    ("Vergaderruimte",        "Dun tapijt",                "Gipsplaat + isolatie",     "Gipsplafond (dicht)",               "Vergaderruimte",                "4. Losse tapijten",         0.6),
    ("Restaurant / cafe",     "Harde vloer (beton/tegels)", "Stucwerk",                "Open plafond (installaties)",       "Restaurant / horeca",           "1. Niks extra",             1.0),
    ("Lobby / receptie",      "Harde vloer (beton/tegels)", "Overwegend glas (~75%)",  "Betonplafond",                      "Lounge / gestoffeerd",          "2. Alleen planten",         1.0),
    ("Klaslokaal",            "Linoleum / vinyl / PVC",    "Gipsplaat (standaard)",    "Systeemplafond (oud/dun)",          "Onderwijs / school",            "1. Niks extra",             0.8),
    ("Zorginstelling",        "Linoleum / vinyl / PVC",    "Gipsplaat + isolatie",     "Systeemplafond (standaard)",        "Zorg / medisch",                "3. Dunne gordijnen",        0.8),
    ("Retail / winkel",       "Harde vloer (beton/tegels)", "Glazen systeemwand",      "Open plafond (installaties)",       "Retail / winkel",               "1. Niks extra",             1.0),
]

# Peutz absorptiedata (EN-ISO 354:2003, rapport A 3432-1-RA)
PEUTZ_V5 = {
    "a_125Hz": 0.25, "a_250Hz": 0.65, "a_500Hz": 0.99,
    "a_1000Hz": 0.99, "a_2000Hz": 0.97, "a_4000Hz": 0.93,
}

PEUTZ_FLAT = [
    [125, .02, .04, .09, .09, .26],
    [250, .03, .12, .31, .28, .73],
    [500, .12, .40, .71, .69, .99],
    [1000, .33, .71, .97, .98, .99],
    [2000, .62, .87, .94, .99, .99],
    [4000, .81, .86, .81, .97, .99],
]
PEUTZ_FLAT_AW = ["αw", "0.20(H)", "0.45(L)", "0.70", "0.65", "1.00"]
PEUTZ_FLAT_CL = ["Klasse", "E", "D", "C", "C", "A"]

PEUTZ_DIV = [
    [125, .19, .19, .17, .24, .25],
    [250, .25, .40, .41, .52, .65],
    [500, .42, .83, .86, .99, .99],
    [1000, .71, .93, .91, .99, .99],
    [2000, .86, .88, .87, .94, .97],
    [4000, .86, .82, .79, .89, .93],
]
PEUTZ_DIV_AW = ["αw", "0.45(L)", "0.70", "0.70", "0.85", "0.90"]
PEUTZ_DIV_CL = ["Klasse", "D", "C", "C", "B", "A"]


# ══════════════════════════════════════════════════════════════
# PRODUCT DATA PROCESSING
# ══════════════════════════════════════════════════════════════

def calc_area(row):
    """Calculate acoustic surface area from product dimensions."""
    if pd.notna(row["Acoustic_Surface_m2"]):
        return row["Acoustic_Surface_m2"], False
    fam = row["Product_Family"]
    var = str(row["Product_Variant"])
    d = row["Diameter_mm"]
    a = None
    if fam == "Float" and pd.notna(d):
        a = 2 * math.pi * (d / 2000) ** 2
    elif fam == "Float" and "Oval" in var:
        a = 2 * math.pi * 0.5 * 1.0 if "2000" in var else 2 * math.pi * 0.4 * 0.6
    elif fam == "Float" and "Rectangle" in var:
        if "1200x2400" in var:
            a = 2 * 1.2 * 2.4
        elif "600x2400" in var:
            a = 2 * 0.6 * 2.4
        elif "1200x1200" in var:
            a = 2 * 1.2 * 1.2
    elif fam == "Line":
        a = 2 * 0.15 * 1.29 if "1290" in var else 2 * 0.15 * 2.52
    elif fam == "Bold" and "1200x1200" in var:
        a = 2 * 1.2 * 1.2
    elif fam == "Cone" and pd.notna(d):
        r_m = d / 2000
        a = math.pi * r_m * math.sqrt(r_m ** 2 + 0.35 ** 2)
    elif fam == "Macaron" and "pendant" in var and pd.notna(d):
        a = 0.6 * 4 * math.pi * (d / 2000) ** 2
    elif fam == "Macaron" and "wall" in var and pd.notna(d):
        a = math.pi * (d / 2000) ** 2
    elif fam == "Edge":
        a = 2 * 0.6 * 1.2 if "XL" in var else 2 * 0.4 * 0.8
    return (round(a, 2), True) if a else (None, True)


def load_products(csv_path):
    """Load and process product data from CSV."""
    cat = pd.read_csv(csv_path, sep=";", encoding="utf-8-sig")

    # Calculate areas
    cat["area_calculated"] = False
    for idx, row in cat.iterrows():
        area, calc = calc_area(row)
        cat.at[idx, "Acoustic_Surface_m2"] = area
        cat.at[idx, "area_calculated"] = calc

    # Apply Peutz V5 data to Float acoustic variants
    fmask = (cat["Product_Family"] == "Float") & cat["Product_Variant"].str.contains(
        "acoustic", case=False, na=False
    )
    for col, val in PEUTZ_V5.items():
        cat.loc[fmask, col] = val

    # Calculate derived columns
    cat["Display_Name"] = cat["Product_Family"] + " " + cat["Product_Variant"]
    cat["Speech_Avg"] = cat[["a_500Hz", "a_1000Hz", "a_2000Hz"]].mean(axis=1).round(2)
    cat["Diff_Factor"] = 1.00
    pendant_mask = cat["Mounting_Type"].str.contains(
        "pendant|hang|free", case=False, na=False
    )
    cat.loc[pendant_mask, "Diff_Factor"] = 1.12
    cat["Aeq_unit"] = (
        cat["Acoustic_Surface_m2"] * cat["Speech_Avg"] * cat["Diff_Factor"]
    ).round(2)

    return cat


# ══════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════

def write_mat_block(ws, sr, title, hdrs, data, has_alpha=True):
    """Write a material data block and return data start/end rows."""
    ws.cell(row=sr, column=1, value=title).font = sf
    hr = sr + 1
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=hr, column=i, value=h)
    shdr(ws, hr, 1, len(hdrs))

    for i, rd in enumerate(data):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=rd[0]).font = df_
        ws.cell(row=r, column=1).border = bdr
        if has_alpha:
            for j, v in enumerate(rd[1:], 2):
                cl = ws.cell(row=r, column=j, value=v)
                cl.font = df_
                cl.border = bdr
                cl.number_format = "0.00"
            speech = round((rd[3] + rd[4] + rd[5]) / 3, 2)
            cl = ws.cell(row=r, column=8, value=speech)
            cl.font = lbf
            cl.border = bdr
            cl.number_format = "0.00"
        else:
            cl = ws.cell(row=r, column=2, value=rd[1])
            cl.font = df_
            cl.border = bdr
            cl.number_format = "0.000"

    ds = hr + 1
    de = hr + len(data)
    srange(ws, ds, de, 1, len(hdrs))
    return ds, de


def build_materiaaldata(wb):
    """Build the Materiaaldata sheet. Returns row references."""
    ws = wb.active
    ws.title = "Materiaaldata"
    ws.sheet_properties.tabColor = "548235"
    ws["A1"] = "MATERIAALDATA & LOOKUP V4"
    ws["A1"].font = tf
    ws["A2"] = "Bronnen: Acoustic.ua, Sengpiel Audio, Peutz | Pas hier aan om Calculator te wijzigen"
    ws["A2"].font = hf

    fh = ["Materiaal", "α 125", "α 250", "α 500", "α 1000", "α 2000", "α 4000", "Spraak gem."]
    fh2 = ["Omschrijving", "Factor per m² vloer"]

    fl_s, fl_e = write_mat_block(ws, 4, "VLOER", fh, FLOORS)
    wa_s, wa_e = write_mat_block(ws, fl_e + 2, "WANDEN", fh, WALLS)
    pl_s, pl_e = write_mat_block(ws, wa_e + 2, "PLAFOND", fh, CEILINGS)
    in_s, in_e = write_mat_block(ws, pl_e + 2, "INRICHTING", fh2, FURNITURE, False)
    ov_s, ov_e = write_mat_block(ws, in_e + 2, "OVERIGE AKOESTIEK", fh2, OTHER_AC, False)

    # Presets
    pr_tr = ov_e + 2
    ws.cell(row=pr_tr, column=1, value="RUIMTETYPE PRESETS").font = sf
    pr_h = ["Ruimtetype", "Vloer", "Wanden", "Plafond", "Inrichting", "Overige akoestiek", "Norm RT60"]
    pr_hr = pr_tr + 1
    for i, h in enumerate(pr_h, 1):
        ws.cell(row=pr_hr, column=i, value=h)
    shdr(ws, pr_hr, 1, len(pr_h))

    for i, p in enumerate(PRESETS):
        r = pr_hr + 1 + i
        for j, v in enumerate(p):
            cl = ws.cell(row=r, column=j + 1, value=v)
            cl.font = df_
            cl.border = bdr
    pr_s = pr_hr + 1
    pr_e = pr_hr + len(PRESETS)

    # Column widths
    for i, w in enumerate([38, 10, 10, 10, 10, 10, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return {
        "fl": (fl_s, fl_e), "wa": (wa_s, wa_e), "pl": (pl_s, pl_e),
        "in": (in_s, in_e), "ov": (ov_s, ov_e), "pr": (pr_s, pr_e),
    }


def build_productdata(wb, cat):
    """Build the Productdata sheet. Returns product header row and end row."""
    ws = wb.create_sheet("Productdata")
    ws.sheet_properties.tabColor = MED
    ws["A1"] = "LUMENEAR PRODUCTDATA 2026"
    ws["A1"].font = tf
    ws["A2"] = "★ = berekend opp. | Pendant +12% randdiffractie (bronpositie, randeffect, open vormen)"
    ws["A2"].font = Font(size=9, italic=True, color="EA580C", name="Calibri")

    ph = ["Product", "Familie", "αw", "Opp. (m²)", "Bron", "Spraak α", "Diffractie", "Aeq/stuk (m²)", "Mounting"]
    p_hr = 4
    for i, h in enumerate(ph, 1):
        ws.cell(row=p_hr, column=i, value=h)
    shdr(ws, p_hr, 1, len(ph))

    for idx, row in cat.iterrows():
        r = p_hr + 1 + idx
        ws.cell(row=r, column=1, value=row["Display_Name"])
        ws.cell(row=r, column=2, value=row["Product_Family"])
        ws.cell(row=r, column=3, value=row["aw_ISO11654"])
        ws.cell(row=r, column=4, value=row["Acoustic_Surface_m2"])
        ws.cell(row=r, column=5, value="★ Berekend" if row["area_calculated"] else "Catalogus")
        ws.cell(row=r, column=6, value=row["Speech_Avg"])
        ws.cell(row=r, column=7, value=row["Diff_Factor"])
        ws.cell(row=r, column=8, value=row["Aeq_unit"])
        ws.cell(row=r, column=9, value=row["Mounting_Type"])
        if row["area_calculated"]:
            ws.cell(row=r, column=4).fill = orng_fill
            ws.cell(row=r, column=5).fill = orng_fill
        if row["Diff_Factor"] > 1:
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="DBEAFE")
            ws.cell(row=r, column=7).number_format = "0%"
            ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="DBEAFE")

    p_end = p_hr + len(cat)
    srange(ws, p_hr + 1, p_end, 1, len(ph))
    ws.auto_filter.ref = f"A{p_hr}:{get_column_letter(len(ph))}{p_end}"
    for i, w in enumerate([34, 12, 8, 10, 14, 10, 10, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{p_hr + 1}"

    return p_hr, p_end


def build_calculator(wb, refs, p_hr, p_end):
    """Build the Calculator sheet."""
    ws = wb.create_sheet("Calculator")
    ws.sheet_properties.tabColor = "EA580C"
    wb.move_sheet("Calculator", offset=-2)

    fl_s, fl_e = refs["fl"]
    wa_s, wa_e = refs["wa"]
    pl_s, pl_e = refs["pl"]
    in_s, in_e = refs["in"]
    ov_s, ov_e = refs["ov"]
    pr_s, pr_e = refs["pr"]

    # ① HEADER
    ws.merge_cells("A1:F1")
    ws["A1"] = "LUMENEAR AKOESTISCHE CALCULATOR"
    ws["A1"].font = Font(bold=True, size=18, color=DARK, name="Calibri")
    ws["A2"] = "Indicatieve berekening nagalmtijd | Sabine | Spraak 500-2000 Hz"
    ws["A2"].font = Font(size=10, italic=True, color="6B7280", name="Calibri")
    ws["A3"] = "▸ Vul de blauwe velden in"
    ws["A3"].font = Font(size=10, color=INPUT_C, name="Calibri")

    # ① RUIMTE
    ws["A4"] = "① RUIMTE"
    ws["A4"].font = sf
    dims = [("Lengte (m)", 12), ("Breedte (m)", 8), ("Hoogte (m)", 3)]
    for i, (lab, val) in enumerate(dims):
        r = 5 + i
        ws.cell(row=r, column=1, value=lab).font = lbf
        ws.cell(row=r, column=1).border = bdr
        c = ws.cell(row=r, column=2, value=val)
        c.font = inp_f; c.fill = inp_fill; c.border = bdr; c.number_format = "0.0"

    calcs = [
        ("Volume", "=B5*B6*B7", "m³"),
        ("Vloeroppervlak", "=B5*B6", "m²"),
        ("Plafondoppervlak", "=B5*B6", "m²"),
        ("Wandoppervlak", "=2*(B5*B7)+2*(B6*B7)", "m²"),
    ]
    for i, (lab, formula, unit) in enumerate(calcs):
        r = 5 + i
        ws.cell(row=r, column=4, value=lab).font = lf
        ws.cell(row=r, column=4).border = bdr
        ws.cell(row=r, column=5, value=formula).font = lbf
        ws.cell(row=r, column=5).border = bdr
        ws.cell(row=r, column=5).number_format = "0.0"
        ws.cell(row=r, column=6, value=unit).font = hf

    # ② RUIMTE & MATERIALEN
    ws["A9"] = "② RUIMTE & MATERIALEN"
    ws["A9"].font = sf

    # Preset dropdown
    ws.cell(row=10, column=1, value="Type ruimte").font = lbf
    ws.cell(row=10, column=1).border = bdr
    pc = ws.cell(row=10, column=2, value="(zelf invullen)")
    pc.font = inp_f; pc.fill = inp_fill; pc.border = bdr
    dv_pr = DataValidation(type="list", formula1=f"Materiaaldata!$A${pr_s}:$A${pr_e}", allow_blank=False)
    ws.add_data_validation(dv_pr)
    dv_pr.add(pc)

    # C10: norm
    ws.cell(row=10, column=3, value=f'=IF(B10="(zelf invullen)","",VLOOKUP(B10,Materiaaldata!$A${pr_s}:$G${pr_e},7,FALSE))')
    ws.cell(row=10, column=3).font = Font(size=10, color="EA580C", bold=True, name="Calibri")
    ws.cell(row=10, column=3).number_format = '0.0"s norm"'

    # D10: hint
    ws.cell(row=10, column=4, value='=IF(B10="(zelf invullen)","","← Richtlijn")')
    ws.cell(row=10, column=4).font = hf

    # Sub-header
    for i, h in enumerate(["", "Materiaal", "Absorptie (m²)", "Handmatig (m²)", "Aanbevolen"], 1):
        cl = ws.cell(row=12, column=i, value=h)
        cl.font = Font(bold=True, size=9, color="6B7280", name="Calibri")
        cl.fill = sub_fill; cl.border = bdr

    # Material rows
    mat_config = [
        (13, "Vloer",             f"Materiaaldata!$A${fl_s}:$A${fl_e}", f"Materiaaldata!$A${fl_s}:$H${fl_e}", 8, "E6", 2),
        (14, "Wanden",            f"Materiaaldata!$A${wa_s}:$A${wa_e}", f"Materiaaldata!$A${wa_s}:$H${wa_e}", 8, "E8", 3),
        (15, "Plafond",           f"Materiaaldata!$A${pl_s}:$A${pl_e}", f"Materiaaldata!$A${pl_s}:$H${pl_e}", 8, "E7", 4),
        (16, "Inrichting",        f"Materiaaldata!$A${in_s}:$A${in_e}", f"Materiaaldata!$A${in_s}:$B${in_e}", 2, "E6", 5),
        (17, "Overige akoestiek", f"Materiaaldata!$A${ov_s}:$A${ov_e}", f"Materiaaldata!$A${ov_s}:$B${ov_e}", 2, "E6", 6),
    ]
    for r, label, dv_rng, vl_rng, vcol, area_ref, pcol in mat_config:
        ws.cell(row=r, column=1, value=label).font = lbf
        ws.cell(row=r, column=1).border = bdr

        default = "Weet ik niet" if r <= 16 else "1. Niks extra"
        if r == 14:
            default = "Stucwerk"  # wanden has no "Weet ik niet"

        c = ws.cell(row=r, column=2, value=default)
        c.font = inp_f; c.fill = inp_fill; c.border = bdr
        dv = DataValidation(type="list", formula1=dv_rng, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(c)

        ws.cell(row=r, column=3, value=f'=IF(D{r}<>"",D{r},VLOOKUP(B{r},{vl_rng},{vcol},FALSE)*{area_ref})')
        ws.cell(row=r, column=3).font = lbf
        ws.cell(row=r, column=3).border = bdr
        ws.cell(row=r, column=3).number_format = "0.0"

        ws.cell(row=r, column=4).font = inp_f
        ws.cell(row=r, column=4).fill = inp_fill
        ws.cell(row=r, column=4).border = bdr
        ws.cell(row=r, column=4).number_format = "0.0"

        ws.cell(row=r, column=5, value=f'=IF(B10="(zelf invullen)","",VLOOKUP(B10,Materiaaldata!$A${pr_s}:$G${pr_e},{pcol},FALSE))')
        ws.cell(row=r, column=5).font = Font(size=9, italic=True, color="059669", name="Calibri")

    ws.cell(row=18, column=4, value="← optioneel: overschrijf met m²").font = hf

    # ③ LUMENEAR PRODUCTEN
    ws["A19"] = "③ LUMENEAR PRODUCTEN"
    ws["A19"].font = sf
    for i, h in enumerate(["Product", "Aantal", "m²/stuk", "αw", "Spraak α", "Aeq totaal (m²)"], 1):
        cl = ws.cell(row=20, column=i, value=h)
        cl.font = Font(bold=True, size=9, color="6B7280", name="Calibri")
        cl.fill = sub_fill; cl.border = bdr

    PS, PE = 21, 26
    prod_ref = f"Productdata!$A${p_hr + 1}:$A${p_end}"
    for r in range(PS, PE + 1):
        c = ws.cell(row=r, column=1)
        c.font = inp_f; c.fill = inp_fill; c.border = bdr
        dv_p = DataValidation(type="list", formula1=prod_ref, allow_blank=True)
        ws.add_data_validation(dv_p)
        dv_p.add(c)

        ws.cell(row=r, column=2).font = inp_f
        ws.cell(row=r, column=2).fill = inp_fill
        ws.cell(row=r, column=2).border = bdr

        # Lookups
        pr = f"Productdata!${{}}${p_hr + 1}:${{}}${p_end}"
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",INDEX(Productdata!$D${p_hr+1}:$D${p_end},MATCH(A{r},Productdata!$A${p_hr+1}:$A${p_end},0)))')
        ws.cell(row=r, column=3).font = df_; ws.cell(row=r, column=3).border = bdr; ws.cell(row=r, column=3).number_format = "0.00"
        ws.cell(row=r, column=4, value=f'=IF(A{r}="","",INDEX(Productdata!$C${p_hr+1}:$C${p_end},MATCH(A{r},Productdata!$A${p_hr+1}:$A${p_end},0)))')
        ws.cell(row=r, column=4).font = df_; ws.cell(row=r, column=4).border = bdr; ws.cell(row=r, column=4).number_format = "0.00"
        ws.cell(row=r, column=5, value=f'=IF(A{r}="","",INDEX(Productdata!$F${p_hr+1}:$F${p_end},MATCH(A{r},Productdata!$A${p_hr+1}:$A${p_end},0)))')
        ws.cell(row=r, column=5).font = df_; ws.cell(row=r, column=5).border = bdr; ws.cell(row=r, column=5).number_format = "0.00"
        ws.cell(row=r, column=6, value=f'=IF(OR(A{r}="",B{r}=""),0,B{r}*INDEX(Productdata!$H${p_hr+1}:$H${p_end},MATCH(A{r},Productdata!$A${p_hr+1}:$A${p_end},0)))')
        ws.cell(row=r, column=6).font = Font(bold=True, size=10, color="EA580C", name="Calibri")
        ws.cell(row=r, column=6).border = bdr; ws.cell(row=r, column=6).number_format = "0.00"

    # Default products
    ws.cell(row=21, column=1, value="Halo light 700"); ws.cell(row=21, column=2, value=0)
    ws.cell(row=22, column=1, value="Halo light 1400"); ws.cell(row=22, column=2, value=1)
    ws.cell(row=23, column=1, value="Float acoustic Rectangle 1200x2400"); ws.cell(row=23, column=2, value=5)
    ws.cell(row=24, column=1, value="Toad 1500"); ws.cell(row=24, column=2, value=1)

    # ④ RESULTAAT
    ws["A28"] = "④ RESULTAAT"
    ws["A28"].font = Font(bold=True, size=14, color="EA580C", name="Calibri")

    ws.cell(row=29, column=1, value="Ruimte-absorptie (spraak)").font = lf
    ws.cell(row=29, column=1).border = bdr
    ws.cell(row=29, column=2, value="=C13+C14+C15+C16+C17").font = lbf
    ws.cell(row=29, column=2).border = bdr; ws.cell(row=29, column=2).number_format = "0.0"
    ws.cell(row=29, column=3, value="m² Aeq").font = hf

    ws.cell(row=30, column=1, value="Lumenear absorptie (spraak)").font = lf
    ws.cell(row=30, column=1).fill = orng_fill; ws.cell(row=30, column=1).border = bdr
    ws.cell(row=30, column=2, value=f"=SUM(F{PS}:F{PE})")
    ws.cell(row=30, column=2).font = Font(bold=True, size=11, color="EA580C", name="Calibri")
    ws.cell(row=30, column=2).fill = orng_fill; ws.cell(row=30, column=2).border = bdr
    ws.cell(row=30, column=2).number_format = "0.0"
    ws.cell(row=30, column=3, value="m² Aeq").font = hf

    # RT60
    ws.cell(row=32, column=1, value="Nagalmtijd ZONDER Lumenear").font = Font(bold=True, size=13, color="DC2626", name="Calibri")
    ws.cell(row=32, column=1).border = bdr2
    ws.cell(row=32, column=2, value='=IF(B29<=0,"n.v.t.",MIN(0.161*E5/B29,6))')
    ws.cell(row=32, column=2).font = Font(bold=True, size=20, color="DC2626", name="Calibri")
    ws.cell(row=32, column=2).border = bdr2; ws.cell(row=32, column=2).number_format = "0.0"
    ws.cell(row=32, column=3, value="sec").font = lf

    ws.cell(row=33, column=1, value="Nagalmtijd MET Lumenear").font = Font(bold=True, size=13, color="059669", name="Calibri")
    ws.cell(row=33, column=1).fill = grn_fill; ws.cell(row=33, column=1).border = bdr2
    ws.cell(row=33, column=2, value='=IF((B29+B30)<=0,"n.v.t.",MIN(0.161*E5/(B29+B30),6))')
    ws.cell(row=33, column=2).font = Font(bold=True, size=20, color="059669", name="Calibri")
    ws.cell(row=33, column=2).fill = grn_fill; ws.cell(row=33, column=2).border = bdr2
    ws.cell(row=33, column=2).number_format = "0.0"
    ws.cell(row=33, column=3, value="sec").font = lf

    # Verbetering
    ws.cell(row=34, column=1, value="Verbetering").font = Font(bold=True, size=12, color=DARK, name="Calibri")
    ws.cell(row=34, column=1).border = bdr
    ws.cell(row=34, column=2, value='=IF(OR(B32="n.v.t.",B33="n.v.t."),"",1-B33/B32)')
    ws.cell(row=34, column=2).font = Font(bold=True, size=18, color=DARK, name="Calibri")
    ws.cell(row=34, column=2).border = bdr; ws.cell(row=34, column=2).number_format = "0%"

    # Rating
    N = 'IF(C10="",0.8,C10)'
    ws.cell(row=35, column=1, value=(
        f'=IF(B33="n.v.t.","",'
        f'IF(B33<={N},'
        f'"⭐ Uitstekend - Perfecte spraakverstaanbaarheid",'
        f'IF(B33<={N}*1.15,'
        f'"✔️ Prima - Vaak werkbaar",'
        f'IF(B33<={N}*1.3,'
        f'"⚠️ Redelijk - Lichte nagalm, ruimte voor optimalisatie",'
        f'IF(B33<={N}*1.6,'
        f'"🟠 Matig - Norm overschreden, extra absorptie geadviseerd",'
        f'IF(B33<={N}*2,'
        f'"🔴 Onvoldoende - Storende nagalm, actie vereist",'
        f'"🔴 Zeer slecht - Akoestisch onwerkbaar"))))))'
    ))
    ws["A35"].font = Font(bold=True, size=11, name="Calibri")
    ws.merge_cells("A35:F35")

    # Conditional formatting for rating colors
    cf_rules = [
        (f'AND(B33<>"n.v.t.",B33<={N})',       "166534"),
        (f'AND(B33<>"n.v.t.",B33<={N}*1.15)',   "22A06B"),
        (f'AND(B33<>"n.v.t.",B33<={N}*1.3)',    "9A6700"),
        (f'AND(B33<>"n.v.t.",B33<={N}*1.6)',    "EA580C"),
        (f'AND(B33<>"n.v.t.",B33<={N}*2)',      "DC2626"),
        (f'AND(B33<>"n.v.t.",B33>{N}*2)',       "991B1B"),
    ]
    for formula, color in cf_rules:
        ws.conditional_formatting.add(
            "A35",
            FormulaRule(
                formula=[formula],
                font=Font(bold=True, size=11, color=color, name="Calibri"),
                stopIfTrue=True,
            ),
        )

    # A36: preset description
    ws.cell(row=36, column=1, value=(
        f'=IF(B10="(zelf invullen)","Gekozen: zelf invullen -> richtlijn norm = handmatig",'
        f'"Gekozen: "&LOWER(B10)&" -> richtlijn norm = "&VLOOKUP(B10,Materiaaldata!$A${pr_s}:$G${pr_e},7,FALSE)&"s")'
    ))
    ws["A36"].font = Font(size=9, italic=True, color="6B7280", name="Calibri")

    # Chart
    ws["H1"] = "Grafiekdata"
    ws["H1"].font = Font(size=8, color="D1D5DB", name="Calibri")
    ws["H2"] = "Zonder"; ws["H3"] = "Met Lumenear"
    ws["I2"] = "=B32"; ws["I3"] = "=B33"

    chart = BarChart()
    chart.type = "col"; chart.style = 10
    chart.title = "Nagalmtijd voor / na Lumenear"
    chart.y_axis.title = "RT60 (seconden)"
    chart.y_axis.scaling.min = 0
    data = Reference(ws, min_col=9, min_row=1, max_row=3)
    cats = Reference(ws, min_col=8, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12; chart.height = 10
    s = chart.series[0]
    pt0 = DataPoint(idx=0); pt0.graphicalProperties.solidFill = "DC2626"
    pt1 = DataPoint(idx=1); pt1.graphicalProperties.solidFill = "059669"
    s.data_points = [pt0, pt1]
    s.graphicalProperties.line.noFill = True
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.numFmt = '0.0"s"'
    ws.add_chart(chart, "D37")

    ws.cell(row=38, column=1, value="Richtwaarden:").font = Font(bold=True, size=9, color="6B7280", name="Calibri")
    ws.cell(row=39, column=1, value="⚠ Indicatieve berekening.  Pendant producten incl. +12% randdiffractie-bonus.")
    ws["A39"].font = Font(size=9, italic=True, color="DC2626", name="Calibri")

    # Column widths
    for i, w in enumerate([32, 16, 14, 14, 28, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 8
    ws.freeze_panes = "A4"


def build_peutz(wb):
    """Build the Peutz Testdata sheet."""
    ws = wb.create_sheet("Peutz Testdata")
    ws.sheet_properties.tabColor = "4472C4"
    ws["A1"] = "PEUTZ ABSORPTIEDATA — EN-ISO 354:2003"
    ws["A1"].font = tf
    ws["A2"] = "Rapport A 3432-1-RA | 17 mei 2018"
    ws["A2"].font = hf

    # Flat absorbers
    ws["A4"] = "VLAKKE ABSORBERS"
    ws["A4"].font = sf
    for i, h in enumerate(["Hz", "V1: 1x9mm", "V2: 2x9mm", "V3: 3x9mm", "V4: 9mm+20Mét", "V5: 9mm+40Mét"], 1):
        ws.cell(row=5, column=i, value=h)
    shdr(ws, 5, 1, 6)
    for i, row in enumerate(PEUTZ_FLAT):
        for j, v in enumerate(row):
            ws.cell(row=6 + i, column=1 + j, value=v)
    srange(ws, 6, 11, 1, 6)

    for i, row in enumerate([PEUTZ_FLAT_AW, PEUTZ_FLAT_CL]):
        for j, v in enumerate(row):
            cl = ws.cell(row=13 + i, column=1 + j, value=v)
            cl.font = lbf; cl.fill = sub_fill; cl.border = bdr

    # Room dividers
    ws["A17"] = "ROOM-DIVIDERS (2-zijdig)"
    ws["A17"].font = sf
    for i, h in enumerate(["Hz", "V6: 1x9mm", "V7: 3x9mm 1el", "V8: 3x9mm 2el", "V9: 2x9+20Mét", "V10: 2x9+40Mét"], 1):
        ws.cell(row=18, column=i, value=h)
    shdr(ws, 18, 1, 6)
    for i, row in enumerate(PEUTZ_DIV):
        for j, v in enumerate(row):
            ws.cell(row=19 + i, column=1 + j, value=v)
    srange(ws, 19, 24, 1, 6)

    for i, row in enumerate([PEUTZ_DIV_AW, PEUTZ_DIV_CL]):
        for j, v in enumerate(row):
            cl = ws.cell(row=26 + i, column=1 + j, value=v)
            cl.font = lbf; cl.fill = sub_fill; cl.border = bdr

    # Conditional formatting for Peutz data
    for col in ["B", "C", "D", "E", "F"]:
        for rng in [f"{col}6:{col}11", f"{col}19:{col}24"]:
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], fill=grn_fill))
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.4", "0.79"], fill=PatternFill("solid", fgColor=SPEECH_BG)))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.4"], fill=red_fill))

    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 16


# ══════════════════════════════════════════════════════════════
# MAIN BUILD
# ══════════════════════════════════════════════════════════════

def build(csv_path, output_path):
    """Build the complete calculator."""
    print(f"📂 CSV: {csv_path}")
    print(f"📁 Output: {output_path}")

    # Validate dropdown lengths
    for name, items in [("Vloer", FLOORS), ("Wanden", WALLS), ("Plafond", CEILINGS),
                        ("Inrichting", FURNITURE), ("Overige", OTHER_AC)]:
        length = len(",".join([i[0] for i in items]))
        status = "✓" if length < 255 else "🔴 OVER 255!"
        print(f"  Dropdown {name}: {length} tekens {status}")
        if length >= 255:
            print(f"  WAARSCHUWING: {name} dropdown te lang voor Excel validatie!")

    # Load products
    print(f"\n📊 Laden productdata...")
    cat = load_products(csv_path)
    print(f"  {len(cat)} producten geladen")
    print(f"  {cat['area_calculated'].sum()} oppervlaktes berekend (★)")
    print(f"  {(cat['Diff_Factor'] > 1).sum()} pendant producten (+12% diffractie)")

    # Build workbook
    wb = Workbook()

    print(f"\n🔨 Bouwen Materiaaldata...")
    refs = build_materiaaldata(wb)
    print(f"  Rijen: vloer={refs['fl']} wanden={refs['wa']} plafond={refs['pl']}")
    print(f"         inrichting={refs['in']} overige={refs['ov']} presets={refs['pr']}")

    print(f"🔨 Bouwen Productdata...")
    p_hr, p_end = build_productdata(wb, cat)
    print(f"  Producten: rij {p_hr+1} t/m {p_end}")

    print(f"🔨 Bouwen Calculator...")
    build_calculator(wb, refs, p_hr, p_end)

    print(f"🔨 Bouwen Peutz Testdata...")
    build_peutz(wb)

    # Save
    wb.save(output_path)
    size_kb = os.path.getsize(output_path) / 1024
    print(f"\n✅ Opgeslagen: {output_path} ({size_kb:.0f} KB)")
    print(f"   4 sheets: Calculator | Materiaaldata | Productdata | Peutz Testdata")
    print(f"   89 producten | 12 wand-opties | 12 overige akoestiek | 9 presets")
    print(f"   6-level rating met conditional formatting")


def main():
    parser = argparse.ArgumentParser(description="Bouw Lumenear Akoestische Calculator V4")
    parser.add_argument("--csv", default="lumenear_2026_acoustic_data.csv",
                        help="Pad naar product CSV (default: lumenear_2026_acoustic_data.csv)")
    parser.add_argument("--output", default="lumenear_acoustic_calculator_v4.xlsx",
                        help="Output pad (default: lumenear_acoustic_calculator_v4.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.csv):
        print(f"FOUT: CSV niet gevonden: {args.csv}")
        print(f"Zorg dat lumenear_2026_acoustic_data.csv in dezelfde map staat.")
        sys.exit(1)

    build(args.csv, args.output)


if __name__ == "__main__":
    main()
